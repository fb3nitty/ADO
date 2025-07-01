#!/usr/bin/env python3
"""
Azure DevOps Permission Matrix Excel Template Generator
Creates a comprehensive Excel workbook with multiple worksheets for permission management
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

def create_permission_matrix_excel():
    """Create the comprehensive Azure DevOps Permission Matrix Excel template"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Create Summary Dashboard sheet
    summary_ws = wb.create_sheet("Summary Dashboard")
    create_summary_dashboard(summary_ws, header_font, header_fill, subheader_fill, border, center_align)
    
    # Create Organization Permissions sheet
    org_ws = wb.create_sheet("Organization Permissions")
    create_organization_permissions(org_ws, header_font, header_fill, subheader_fill, border, center_align)
    
    # Create Project Permissions sheet
    project_ws = wb.create_sheet("Project Permissions")
    create_project_permissions(project_ws, header_font, header_fill, subheader_fill, border, center_align)
    
    # Create Repository Permissions sheet
    repo_ws = wb.create_sheet("Repository Permissions")
    create_repository_permissions(repo_ws, header_font, header_fill, subheader_fill, border, center_align)
    
    # Create Pipeline Permissions sheet
    pipeline_ws = wb.create_sheet("Pipeline Permissions")
    create_pipeline_permissions(pipeline_ws, header_font, header_fill, subheader_fill, border, center_align)
    
    # Create Instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    create_instructions_sheet(instructions_ws, header_font, header_fill, subheader_fill, border)
    
    # Set Summary Dashboard as active sheet
    wb.active = summary_ws
    
    # Save workbook
    output_path = "/home/ubuntu/azure_devops_toolkit/templates/permission_matrix.xlsx"
    wb.save(output_path)
    print(f"Excel template created successfully: {output_path}")

def create_summary_dashboard(ws, header_font, header_fill, subheader_fill, border, center_align):
    """Create the Summary Dashboard worksheet"""
    
    # Title
    ws['A1'] = "Azure DevOps Permission Matrix - Summary Dashboard"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    
    # Organization Info section
    ws['A3'] = "Organization Information"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:D3')
    
    org_info = [
        ["Organization Name:", ""],
        ["Organization URL:", ""],
        ["Last Audit Date:", ""],
        ["Audited By:", ""],
        ["Next Review Date:", ""]
    ]
    
    for i, (label, value) in enumerate(org_info, start=4):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # Permission Summary section
    ws['A10'] = "Permission Summary"
    ws['A10'].font = header_font
    ws['A10'].fill = header_fill
    ws.merge_cells('A10:H10')
    
    summary_headers = ["Category", "Total Users", "Admin Users", "Contributors", "Readers", "External Users", "Service Accounts", "Risk Level"]
    for i, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=11, column=i, value=header)
        cell.font = header_font
        cell.fill = subheader_fill
        cell.border = border
        cell.alignment = center_align
    
    summary_data = [
        ["Organization Level", "=COUNTA('Organization Permissions'!B:B)-1", "=COUNTIF('Organization Permissions'!D:D,\"*Admin*\")", "", "", "", "", "High"],
        ["Project Level", "=COUNTA('Project Permissions'!B:B)-1", "=COUNTIF('Project Permissions'!D:D,\"*Admin*\")", "=COUNTIF('Project Permissions'!D:D,\"*Contributor*\")", "=COUNTIF('Project Permissions'!D:D,\"*Reader*\")", "", "", "Medium"],
        ["Repository Level", "=COUNTA('Repository Permissions'!B:B)-1", "", "", "", "", "", "Medium"],
        ["Pipeline Level", "=COUNTA('Pipeline Permissions'!B:B)-1", "", "", "", "", "", "Medium"]
    ]
    
    for i, row_data in enumerate(summary_data, start=12):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = border
            if j == 8:  # Risk Level column
                if value == "High":
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif value == "Medium":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Key Findings section
    ws['A17'] = "Key Findings & Recommendations"
    ws['A17'].font = header_font
    ws['A17'].fill = header_fill
    ws.merge_cells('A17:H17')
    
    findings = [
        "• Review organization-level administrators - limit to essential personnel only",
        "• Ensure all repositories have branch protection policies enabled",
        "• Verify service connections use minimal required permissions",
        "• Check for inactive users and remove unnecessary access",
        "• Validate external user access is properly scoped and time-limited"
    ]
    
    for i, finding in enumerate(findings, start=18):
        ws[f'A{i}'] = finding
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_organization_permissions(ws, header_font, header_fill, subheader_fill, border, center_align):
    """Create the Organization Permissions worksheet"""
    
    # Title
    ws['A1'] = "Organization-Level Permissions Matrix"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:J1')
    
    # Headers
    headers = [
        "User/Group Name", "Type", "Email", "Security Group", "Access Level", 
        "Last Login", "Status", "Risk Level", "Review Date", "Notes"
    ]
    
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # Sample data
    sample_data = [
        ["John Admin", "User", "john.admin@company.com", "Project Collection Administrators", "Visual Studio", "2024-01-15", "Active", "High", "2024-04-15", "Primary org admin"],
        ["Jane Admin", "User", "jane.admin@company.com", "Project Collection Administrators", "Visual Studio", "2024-01-14", "Active", "High", "2024-04-15", "Backup org admin"],
        ["DevOps Team", "Azure AD Group", "", "Project Collection Build Administrators", "Basic", "", "Active", "Medium", "2024-04-15", "Build automation group"],
        ["Security Reviewers", "Azure AD Group", "", "Custom Security Group", "Basic", "", "Active", "Medium", "2024-04-15", "Security audit access"],
        ["External Contractors", "Azure AD Group", "", "Project Collection Valid Users", "Stakeholder", "", "Active", "Low", "2024-02-15", "Limited external access"]
    ]
    
    for i, row_data in enumerate(sample_data, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = border
            if j == 8:  # Risk Level column
                if value == "High":
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif value == "Medium":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif value == "Low":
                    cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
    
    # Add data validation for dropdowns
    risk_validation = DataValidation(type="list", formula1='"High,Medium,Low"')
    status_validation = DataValidation(type="list", formula1='"Active,Inactive,Suspended"')
    access_validation = DataValidation(type="list", formula1='"Stakeholder,Basic,Basic + Test Plans,Visual Studio Subscriber"')
    
    ws.add_data_validation(risk_validation)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(access_validation)
    
    risk_validation.add(f'H4:H1000')
    status_validation.add(f'G4:G1000')
    access_validation.add(f'E4:E1000')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_project_permissions(ws, header_font, header_fill, subheader_fill, border, center_align):
    """Create the Project Permissions worksheet"""
    
    # Title
    ws['A1'] = "Project-Level Permissions Matrix"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:L1')
    
    # Headers
    headers = [
        "Project Name", "User/Group Name", "Type", "Security Group", "Boards", "Repos", 
        "Pipelines", "Test Plans", "Artifacts", "Access Level", "Review Date", "Notes"
    ]
    
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # Sample data
    sample_data = [
        ["ProjectAlpha", "Development Team", "Azure AD Group", "Contributors", "Edit", "Contribute", "Edit", "Edit", "Edit", "Basic", "2024-04-15", "Main dev team"],
        ["ProjectAlpha", "QA Team", "Azure AD Group", "Contributors", "Edit", "Read", "Edit", "Edit", "Read", "Basic", "2024-04-15", "Quality assurance"],
        ["ProjectAlpha", "Product Owner", "User", "Readers", "Edit", "Read", "Read", "Read", "Read", "Stakeholder", "2024-04-15", "Business stakeholder"],
        ["ProjectAlpha", "Release Manager", "User", "Release Administrators", "Read", "Read", "Administer", "Read", "Read", "Basic", "2024-04-15", "Deployment management"],
        ["ProjectBeta", "Beta Dev Team", "Azure AD Group", "Contributors", "Edit", "Contribute", "Edit", "Edit", "Edit", "Basic", "2024-04-15", "Beta project team"]
    ]
    
    for i, row_data in enumerate(sample_data, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = border
    
    # Add data validation
    permission_validation = DataValidation(type="list", formula1='"None,Read,Edit,Contribute,Administer"')
    access_validation = DataValidation(type="list", formula1='"Stakeholder,Basic,Basic + Test Plans,Visual Studio Subscriber"')
    
    ws.add_data_validation(permission_validation)
    ws.add_data_validation(access_validation)
    
    permission_validation.add(f'E4:I1000')
    access_validation.add(f'J4:J1000')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_repository_permissions(ws, header_font, header_fill, subheader_fill, border, center_align):
    """Create the Repository Permissions worksheet"""
    
    # Title
    ws['A1'] = "Repository Permissions & Branch Policies"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:M1')
    
    # Headers
    headers = [
        "Project", "Repository", "Branch", "User/Group", "Read", "Contribute", "Force Push", 
        "Manage Permissions", "Policy: Min Reviewers", "Policy: Build Validation", 
        "Policy: Comment Resolution", "Policy: Work Item Link", "Notes"
    ]
    
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # Sample data
    sample_data = [
        ["ProjectAlpha", "MainRepo", "main", "Contributors", "Allow", "Allow", "Deny", "Not Set", "2", "Required", "Required", "Required", "Protected main branch"],
        ["ProjectAlpha", "MainRepo", "develop", "Contributors", "Allow", "Allow", "Allow", "Not Set", "1", "Required", "Optional", "Optional", "Development branch"],
        ["ProjectAlpha", "MainRepo", "*", "Readers", "Allow", "Deny", "Deny", "Deny", "", "", "", "", "Read-only access"],
        ["ProjectAlpha", "ConfigRepo", "main", "Release Managers", "Allow", "Allow", "Deny", "Allow", "2", "Required", "Required", "Required", "Configuration management"],
        ["ProjectBeta", "BetaRepo", "main", "Beta Team", "Allow", "Allow", "Deny", "Not Set", "1", "Required", "Required", "Optional", "Beta main branch"]
    ]
    
    for i, row_data in enumerate(sample_data, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = border
            # Highlight security-critical permissions
            if j in [7, 8] and value in ["Allow", "Deny"]:  # Force Push and Manage Permissions
                if value == "Allow":
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif value == "Deny":
                    cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
    
    # Add data validation
    permission_validation = DataValidation(type="list", formula1='"Allow,Deny,Not Set"')
    policy_validation = DataValidation(type="list", formula1='"Required,Optional,Disabled"')
    
    ws.add_data_validation(permission_validation)
    ws.add_data_validation(policy_validation)
    
    permission_validation.add(f'E4:H1000')
    policy_validation.add(f'J4:L1000')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_pipeline_permissions(ws, header_font, header_fill, subheader_fill, border, center_align):
    """Create the Pipeline Permissions worksheet"""
    
    # Title
    ws['A1'] = "Pipeline Permissions & Security"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:L1')
    
    # Headers
    headers = [
        "Project", "Pipeline Name", "Type", "User/Group", "View", "Edit", "Queue/Create", 
        "Administer", "Service Connections", "Agent Pool", "Variable Groups", "Notes"
    ]
    
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # Sample data
    sample_data = [
        ["ProjectAlpha", "CI-Build", "Build", "Contributors", "Allow", "Allow", "Allow", "Deny", "Azure-Prod", "Default", "Secrets-Prod", "Main CI pipeline"],
        ["ProjectAlpha", "CD-Production", "Release", "Release Managers", "Allow", "Allow", "Allow", "Allow", "Azure-Prod", "Default", "Secrets-Prod", "Production deployment"],
        ["ProjectAlpha", "CD-Staging", "Release", "Contributors", "Allow", "Allow", "Allow", "Deny", "Azure-Dev", "Default", "Secrets-Dev", "Staging deployment"],
        ["ProjectAlpha", "Security-Scan", "Build", "Security Team", "Allow", "Allow", "Allow", "Allow", "Security-Tools", "Security-Agents", "Security-Config", "Security scanning"],
        ["ProjectBeta", "Beta-CI", "Build", "Beta Team", "Allow", "Allow", "Allow", "Deny", "Azure-Dev", "Default", "Beta-Vars", "Beta CI pipeline"]
    ]
    
    for i, row_data in enumerate(sample_data, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = border
            # Highlight administrative permissions
            if j == 8 and value == "Allow":  # Administer permission
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Add data validation
    permission_validation = DataValidation(type="list", formula1='"Allow,Deny,Not Set"')
    
    ws.add_data_validation(permission_validation)
    permission_validation.add(f'E4:H1000')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_instructions_sheet(ws, header_font, header_fill, subheader_fill, border):
    """Create the Instructions worksheet"""
    
    # Title
    ws['A1'] = "Azure DevOps Permission Matrix - Instructions for IT Administrators"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    
    # Overview section
    ws['A3'] = "Overview"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:F3')
    
    overview_text = [
        "This Excel workbook provides a comprehensive template for managing and auditing Azure DevOps permissions.",
        "Use this template to document current permissions, plan changes, and maintain compliance with security policies.",
        "",
        "The workbook contains the following worksheets:",
        "• Summary Dashboard - High-level overview and key metrics",
        "• Organization Permissions - Organization-level security groups and users",
        "• Project Permissions - Project-level permissions and team assignments",
        "• Repository Permissions - Repository access and branch policies",
        "• Pipeline Permissions - Build/release pipeline security settings",
        "• Instructions - This sheet with usage guidance"
    ]
    
    for i, text in enumerate(overview_text, start=4):
        ws[f'A{i}'] = text
    
    # Usage Instructions
    ws['A15'] = "Usage Instructions"
    ws['A15'].font = header_font
    ws['A15'].fill = header_fill
    ws.merge_cells('A15:F15')
    
    instructions = [
        "",
        "1. INITIAL SETUP:",
        "   • Fill in organization information on the Summary Dashboard",
        "   • Run the PowerShell audit scripts to extract current permissions",
        "   • Import the CSV data into the appropriate worksheets",
        "",
        "2. PERMISSION DOCUMENTATION:",
        "   • Use the dropdown menus for consistent data entry",
        "   • Color coding: Red = High Risk, Yellow = Medium Risk, Green = Low Risk",
        "   • Update the Review Date column for regular access reviews",
        "",
        "3. REGULAR MAINTENANCE:",
        "   • Update the matrix monthly or after significant changes",
        "   • Use the Notes column to document justifications and exceptions",
        "   • Review high-risk permissions quarterly",
        "",
        "4. COMPLIANCE REPORTING:",
        "   • Use the Summary Dashboard for executive reporting",
        "   • Export individual worksheets for detailed analysis",
        "   • Maintain historical versions for audit trails"
    ]
    
    for i, text in enumerate(instructions, start=16):
        ws[f'A{i}'] = text
    
    # Best Practices
    ws['A33'] = "Security Best Practices"
    ws['A33'].font = header_font
    ws['A33'].fill = header_fill
    ws.merge_cells('A33:F33')
    
    best_practices = [
        "",
        "• PRINCIPLE OF LEAST PRIVILEGE: Grant only minimum required permissions",
        "• REGULAR REVIEWS: Conduct quarterly access reviews and annual certifications",
        "• GROUP-BASED MANAGEMENT: Use Azure AD groups instead of individual assignments",
        "• MFA ENFORCEMENT: Require multi-factor authentication for administrative accounts",
        "• AUDIT LOGGING: Enable and monitor Azure DevOps audit logs",
        "• BRANCH PROTECTION: Implement branch policies on main/master branches",
        "• SERVICE ACCOUNTS: Use managed identities and service principals",
        "• EXTERNAL ACCESS: Limit and regularly review external user permissions",
        "• DOCUMENTATION: Maintain current documentation and change logs",
        "• INCIDENT RESPONSE: Have procedures for permission-related security incidents"
    ]
    
    for i, text in enumerate(best_practices, start=34):
        ws[f'A{i}'] = text
    
    # Contact Information
    ws['A46'] = "Support and Resources"
    ws['A46'].font = header_font
    ws['A46'].fill = header_fill
    ws.merge_cells('A46:F46')
    
    support_info = [
        "",
        "For additional support and resources:",
        "• Azure DevOps Documentation: https://docs.microsoft.com/azure/devops/",
        "• Security Best Practices: https://docs.microsoft.com/azure/devops/organizations/security/",
        "• PowerShell Scripts: Use the included toolkit scripts for automated auditing",
        "• Troubleshooting: Refer to the TROUBLESHOOT.md file in the toolkit documentation"
    ]
    
    for i, text in enumerate(support_info, start=47):
        ws[f'A{i}'] = text
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 80
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15

if __name__ == "__main__":
    create_permission_matrix_excel()
